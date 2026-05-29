#!/usr/bin/env python3
"""Digitalisiert die handgeschriebene Materialbezugsliste -> Excel."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Materialbezug FS26 extra"

# Spalten:
# Nachname | Vorname | Semester | Datum | Stk. | Abrechnung |
# Material Art | Stärke | Grösse
headers = ["Nachname", "Vorname", "Semester", "Datum", "Stk.",
           "Abrechnung", "Material Art", "Stärke", "Grösse"]

# Daten (Ditto-Zeichen aufgelöst, durchgestrichene Werte/Zeilen ignoriert).
# Abrechnung: "Privat" = Material Privatabrechnung, "Institut" = Institutübernahme.
# Alle Einträge der Liste stehen in der Spalte "Privatabrechnung" -> Privat.
rows = [
    # --- Seite 1 ---
    ["Binder", "Lara", "6", "27.05.", 3, "Privat", "GK", "1.5 mm", "55x80"],
    ["Binder", "Lara", "6", "27.05.", 1, "Privat", "GK", "1 mm", "55x80"],
    ["Binder", "Lara", "6", "27.05.", 1, "Privat", "HK", "2.5 mm", "60x90"],
    ["Boxler", "Micha", "2", "27.05.", 2, "Privat", "HK", "2.5", "60x90"],
    ["Demir", "Seyla", "2", "27.05.", 3, "Privat", "HK", "2.5", "60x90"],
    ["Gritter", "Mirja", "2", "27.05.", 1, "Privat", "HK", "2.5", "60x90"],
    ["Kleiber", "Flurina", "2", "27.05.", 2, "Privat", "HK", "2 mm", "55x80"],
    ["Förster", "Tim", "2", "27.05.", 2, "Privat", "GK", "2 mm", "80x110"],
    ["Alboicker", "Alex", "6", "28.05.", 2, "Privat", "GK", "1.5 mm", "80x55"],
    ["Alboicker", "Alex", "6", "28.05.", 4, "Privat", "HK", "2.5", "60x90"],
    ["Metzger", "Maryna", "MA2", "28.05.", 2, "Privat", "GK", "3", "55x80"],
    ["Metzger", "Maryna", "MA2", "28.05.", 1, "Privat", "GK", "1", "55x80"],
    # --- Seite 2 ---
    ["Esposito", "Silvio", "4", "28.05.26", 4, "Privat", "GK", "2 mm", "55x80"],
    ["Esposito", "Silvio", "4", "28.05.26", 4, "Privat", "GK", "2.5 mm", "55x80"],
    ["Abbasov", "Mahmud", "6", "28.05.", 4, "Privat", "HK", "2.5", "60x90"],
    ["Abbasov", "Mahmud", "6", "28.05.", 2, "Privat", "GK", "2.5", "80x80"],
    ["Fijelbaugla", "Martyna", "MA", "28.05.", 3, "Privat", "GK", "1", "55x80"],
    ["Fijelbaugla", "Martyna", "MA", "28.05.", 8, "Privat", "GK", "2", "55x80"],
    ["Fijelbaugla", "Martyna", "MA", "28.05.", 3, "Privat", "GK", "3", "55x80"],
    ["Fijelbaugla", "Martyna", "MA", "28.05.", 3, "Privat", "GK", "2.5", "55x80"],
    ["Fijelbaugla", "Martyna", "MA", "28.05.", 2, "Privat", "GK", "1.5", "55x80"],
    ["Fijelbaugla", "Martyna", "MA", "28.05.", 1, "Privat", "HK", "2.5", "60x90"],
    ["Bartholomäus", "Noah", "6", "28.05.", 1, "Privat", "GK", "1.5", "55x80"],
    ["Bartholomäus", "Noah", "6", "28.05.", 1, "Privat", "GK", "2", "55x80"],
    # --- Seite 3 ---
    ["Boxler", "Micha", "2", "27.05.", 1, "Privat", "HK", "2", "60x90"],
    ["Kleiber", "Flurina", "2", "27.05.", 2, "Privat", "GK", "2", "55x80"],
    ["Wetzel", "David", "3", "28.05.", 1, "Privat", "MDF", "5", "71x75"],
    ["Trüeb", "Raoul", "2", "27.05.", 4, "Privat", "GK", "2", "55x80"],
    ["Trüeb", "Raoul", "2", "27.05.", 2, "Privat", "GK", "2.5", "55x80"],
    ["Trüeb", "Raoul", "2", "27.05.", 1, "Privat", "GK", "1", "55x80"],
    ["Trüeb", "Raoul", "2", "27.05.", 1, "Privat", "GK", "3", "55x80"],
    ["Spaar", "Elia", "2", "27.05.", 3, "Privat", "GK", "2", "55x80"],
    ["Spaar", "Elia", "2", "27.05.", 3, "Privat", "GK", "1.5", "55x80"],
    ["Spaar", "Elia", "2", "27.05.", 3, "Privat", "GK", "3", "55x80"],
    ["Spaar", "Elia", "2", "27.05.", 3, "Privat", "GK", "1", "55x80"],
]

# Styles
header_fill = PatternFill("solid", fgColor="2F5496")
header_font = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")

ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

for r in rows:
    ws.append(r)

# Border + alignment für Datenzeilen
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
    for cell in row:
        cell.border = border
        if cell.column in (3, 4, 5, 6, 7, 8, 9):
            cell.alignment = center

# Spaltenbreiten
widths = [16, 12, 10, 11, 7, 12, 12, 10, 10]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{ws.max_row}"

out = "Materialbezug_IARCH_FS26_extra.xlsx"
wb.save(out)
print(f"Gespeichert: {out}  ({len(rows)} Datenzeilen)")
